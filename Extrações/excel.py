import os
import openpyxl
import settings
from sys import argv
from TeamProjects import *
from Users import *
from project_requests import InvokeREST

def Excel(tipo):
    # TEAM PROJECTS
    if tipo == Projetos(settings.organization):
        print(f'Extração de projetos da organização {settings.organization} sendo realizada.')
        data = Projetos(settings.organization)
        excel = openpyxl.Workbook()
        planilha = excel.active
        planilha.title = "Lista de projetos"
        planilha.cell(column=1, row=1, value="Nome do projeto")
        indexLinhas = 2

        for projeto in data['value']:
            planilha.cell(column=1, row=indexLinhas, value=projeto['name'])
            indexLinhas += 1
        excel.save(f"TeamProjects/{settings.organization}-Projetos.xlsx")


    elif tipo == Repositorios(settings.organization):
        print(f'Extração de projetos da organização {settings.organization} sendo realizada.')
        data = Projetos(settings.organization)
        os.makedirs('TeamProjects/Repos', exist_ok=True)
        for projeto in data['value']:
            projectName = projeto['name']
            print(f'Coletando repositórios do projeto {projectName}')
            excel = openpyxl.Workbook()
            planilha = excel.active
            planilha.title = "Lista de repositorios"
            planilha.cell(column=1, row=1, value="Nome do Repositório")
            indexLinhas = 2
            repos = Repositorios(settings.organization, projectName)

            for repo in repos['value']:
                repoName = repo['name']
                planilha.cell(column=1, row=indexLinhas, value=repoName)
                indexLinhas += 1

            excel.save(f"TeamProjects/Repos/{projectName}_{settings.organization}-ReposAzureDevOps.xlsx")


    elif tipo == Endpoints(settings.organization):
        print(f'Extração de projetos da organização {settings.organization} sendo realizada.')
        data = Projetos(settings.organization)
        os.makedirs('TeamProjects/Endpoints', exist_ok=True)
        for projeto in data['value']:
            projectName = projeto['name']
            print(f'Coletando Endpoints do projeto {projectName}')
            excel = openpyxl.Workbook()
            planilha = excel.active
            planilha.title = "Lista de endpoints"
            planilha.cell(column=1, row=1, value="Endpoint Subscription")
            planilha.cell(column=2, row=1, value="Endpoint Environment")
            planilha.cell(column=3, row=1, value="Endpoint Creator")
            planilha.cell(column=4, row=1, value="Endpoint Description")
            planilha.cell(column=5, row=1, value="Endpoint compartilhado?")
            indexLinhas = 2
            end = Endpoints(settings.organization, projectName)

            for endpoint in end['value']:
                try:
                    subName = endpoint['data']['subscriptionName']
                    endEnv = endpoint['data']['environment']
                    createdBy = endpoint['createdBy']['displayName']
                    desc = endpoint['description']
                    shared = endpoint['isShared']
                except:
                    subName = 'NO INFO'
                    endEnv = 'NO INFO'
                    createdBy = 'NO INFO'
                    desc = 'NO INFO'
                    shared = 'NO INFO'
                planilha.cell(column=1, row=indexLinhas, value=subName)
                planilha.cell(column=2, row=indexLinhas, value=endEnv)
                planilha.cell(column=3, row=indexLinhas, value=createdBy)
                planilha.cell(column=4, row=indexLinhas, value=desc)
                planilha.cell(column=5, row=indexLinhas, value=shared)
                indexLinhas += 1

            excel.save(f"TeamProjects/Endpoints/{projectName}_{settings.organization}-EndpointsAzureDevOps.xlsx")


    elif tipo == SourceProviders(settings.organization):
        print(f'Extração de projetos da organização {settings.organization} sendo realizada.')
        data = Projetos(settings.organization)
        os.makedirs('TeamProjects/SourceProviders', exist_ok=True)
        for projeto in data['value']:
            projectName = projeto['name']
            print(f'Coletando Source Providers do projeto {projectName}')
            excel = openpyxl.Workbook()
            planilha = excel.active
            planilha.title = "Lista de Source Providers"
            planilha.cell(column=1, row=1, value="Source Name")
            indexLinhas = 2
            sp = SourceProviders(settings.organization, projectName)
            for source in sp['value']:
                spName = source['name']
                planilha.cell(column=1, row=indexLinhas, value=spName)
                indexLinhas +=1

            excel.save(f"TeamProjects/SourceProviders/{projectName}_{settings.organization}-SourceProvidersAzureDevOps.xlsx")


    elif tipo == Builds(settings.organization):
        print(f'Extração de projetos da organização {settings.organization} sendo realizada.')
        data = Projetos(settings.organization)
        os.makedirs('BuildsReleases/Builds', exist_ok=True)
        for projeto in data['value']:
            projectName = projeto['name']
            print(f'Coletando Builds realizados no projeto {projectName}')
            excel = openpyxl.Workbook()
            planilha = excel.active
            planilha.title = "Lista de Builds"
            planilha.cell(column=1, row=1, value="Build Definition ID")
            planilha.cell(column=2, row=1, value="Build Definition Name")
            planilha.cell(column=3, row=1, value="Build Number")
            planilha.cell(column=4, row=1, value="Build Repository Name")
            planilha.cell(column=5, row=1, value="Requested By")
            indexLinhas = 2
            builds = Builds(settings.organization, projectName)

            for build in builds['value']:
                buildID = build['definition']['id']
                buildName = build['definition']['name']
                buildNumber = build['id']
                try:
                    repository = build['repository']['name']
                except:
                    repository = "SEM REPOSITORIO"
                requestedBy = build['requestedBy']['displayName']
                planilha.cell(column=1, row=indexLinhas, value=buildID)
                planilha.cell(column=2, row=indexLinhas, value=buildName)
                planilha.cell(column=3, row=indexLinhas, value=buildNumber)
                planilha.cell(column=4, row=indexLinhas, value=repository)
                planilha.cell(column=5, row=indexLinhas, value=requestedBy)
                indexLinhas += 1

            excel.save(f"BuildsReleases/Builds/{projectName}_{settings.organization}-BuildsAzureDevOps.xlsx")


    elif tipo == Releases(settings.organization):
        print(f'Extração de projetos da organização {settings.organization} sendo realizada.')
        data = Projetos(settings.organization)
        os.makedirs('BuildsReleases/Releases', exist_ok=True)
        for projeto in data['value']:
            projectName = projeto['name']
            print(f'Coletando Releases realizadas no projeto {projectName}')
            excel = openpyxl.Workbook()
            planilha = excel.active
            planilha.title = "Lista de Release"
            planilha.cell(column=1, row=1, value="Release Name")
            planilha.cell(column=2, row=1, value="Release ID")
            planilha.cell(column=3, row=1, value="Release Status")
            planilha.cell(column=4, row=1, value="Release Definition Name")
            planilha.cell(column=5, row=1, value="Release Definition ID")
            planilha.cell(column=6, row=1, value="Created By")
            indexLinhas = 2
            releases = Releases(settings.organization, projectName)

            for release in releases['value']:
                releaseName = release['name']
                releaseID = release['id']
                releaseStatus = release['status']
                releaseDefinitionName = release['releaseDefinition']['name']
                releaseDefinitionID = release['releaseDefinition']['id']
                createdBy = release['createdBy']['displayName']
                planilha.cell(column=1, row=indexLinhas, value=releaseName)
                planilha.cell(column=2, row=indexLinhas, value=releaseID)
                planilha.cell(column=3, row=indexLinhas, value=releaseStatus)
                planilha.cell(column=4, row=indexLinhas, value=releaseDefinitionName)
                planilha.cell(column=5, row=indexLinhas, value=releaseDefinitionID)
                planilha.cell(column=6, row=indexLinhas, value=createdBy)
                indexLinhas +=1

            excel.save(f"BuildsReleases/Releases/{projectName}_{settings.organization}-ReleasesAzureDevOps.xlsx")

    # USERS
    elif tipo == Usuarios(settings.organization):
        print(f'Extração de usuários da organização {settings.organization} sendo realizada.')
        data = Usuarios(settings.organization)
        excel = openpyxl.Workbook()
        planilha = excel.active
        planilha.title = "Lista de Usuários"
        planilha.cell(column=1, row=1, value="Username")
        planilha.cell(column=2, row=1, value="Nome do usuário")
        indexLinhas = 2

        for user in data['value']:
            username = user['principalName']
            displayName = user['displayName']
            planilha.cell(column=1, row=indexLinhas, value=username)
            planilha.cell(column=2, row=indexLinhas, value=displayName)
            indexLinhas += 1

        excel.save(f"Users/{settings.organization}-UsuariosAzureDevOps.xlsx")

    elif tipo == UsuariosPorLicenca(settings.organization):
        print(f'Extração de usuários por licença sendo realizada.')
        data = UsuariosPorLicenca(settings.organization)
        excel = openpyxl.Workbook()
        stakeholder = excel.active
        stakeholder.title = "Stakeholder"
        stakeholder.cell(column=1, row=1, value="Username")
        stakeholder.cell(column=2, row=1, value="Nome do usuário")
        basic = excel.create_sheet('Basic')
        basic.cell(column=1, row=1, value="Username")
        basic.cell(column=2, row=1, value="Nome do usuário")
        vStudio = excel.create_sheet('Visual Studio')
        vStudio.cell(column=1, row=1, value="Username")
        vStudio.cell(column=2, row=1, value="Nome do usuário")
        noId = excel.create_sheet('Não identificado')
        noId.cell(column=1, row=1, value="Username")
        noId.cell(column=2, row=1, value="Nome do usuário")
        indexLinhasS = 2
        indexLinhasB = 2
        indexLinhasV = 2
        indexLinhasN = 2
        indexUser = 1
        for usuario in data['members']:
            username = usuario['user']['principalName']
            displayName = usuario['user']['displayName']
            userLicense = usuario['accessLevel']['licenseDisplayName']
            if userLicense == "Basic":
                basic.cell(column=1, row=indexLinhasB, value=username)
                basic.cell(column=2, row=indexLinhasB, value=displayName)
                indexLinhasB += 1
            elif userLicense == "Stakeholder":
                stakeholder.cell(column=1, row=indexLinhasS, value=username)
                stakeholder.cell(column=2, row=indexLinhasS, value=displayName)
                indexLinhasS += 1
            elif userLicense == "Visual Studio Enterprise subscription":
                vStudio.cell(column=1, row=indexLinhasV, value=username)
                vStudio.cell(column=2, row=indexLinhasV, value=displayName)
                indexLinhasV += 1
            else:
                noId.cell(column=1, row=indexLinhasN, value=username)
                noId.cell(column=2, row=indexLinhasN, value=displayName)
                indexLinhasN += 1

            indexUser =+ 1

        excel.save(f"Users/{settings.organization}-UsuariosPorLicencaAzureDevOps.xlsx")
